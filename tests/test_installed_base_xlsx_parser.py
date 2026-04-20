from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.extraction.installed_base_xlsx_parser import (
    detect_template,
    derive_site_token,
    find_header_row,
    is_supported_installed_base_xlsx,
    normalize_header,
    parse_installed_base_workbook,
    remap_corpus_path,
)


def _write_workbook(path: Path, rows: list[list[object]], *, sheet_name: str = "Sheet1") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


def test_normalize_header_aliases():
    assert normalize_header("PART NUMBER") == "part_number"
    assert normalize_header("QTY.") == "quantity"
    assert normalize_header("Actual Serial Number") == "actual_serial_number"


def test_supported_installed_base_xlsx_paths():
    assert is_supported_installed_base_xlsx(r"D:\CorpusTransfr\Site Inventory\Site Inventory Report_Guam.xlsx")
    assert is_supported_installed_base_xlsx(r"D:\CorpusTransfr\Guam\Pre-inventory.xlsx")
    assert is_supported_installed_base_xlsx(r"D:\CorpusTransfr\As-Built\Wake Parts List.xlsx")
    assert not is_supported_installed_base_xlsx(r"D:\CorpusTransfr\Misc\notes.docx")


def test_detect_template_site_inventory():
    headers = ["part_number", "sub_system", "nomenclature", "quantity", "model_number", "serial_number"]
    assert detect_template(headers) == "site_inventory_qpa"


def test_detect_template_inventory_spares():
    headers = ["sub_system", "item_type", "part_number", "serial_number", "quantity"]
    assert detect_template(headers) == "inventory_spares_qty"


def test_detect_template_as_built():
    headers = ["find_no.", "quantity", "part_number", "description", "manufacturer"]
    assert detect_template(headers) == "as_built_parts_list"


def test_derive_site_token_prefers_fallback():
    assert derive_site_token(r"D:\CorpusTransfr\foo\guam\bar.xlsx", "djibouti") == "djibouti"


def test_derive_site_token_from_path():
    assert derive_site_token(r"D:\CorpusTransfr\verified\IGS\Wake Parts List.xlsx") == "wake"
    assert derive_site_token(r"D:\CorpusTransfr\verified\IGS\Pituffik Parts List.xlsx") == "thule"


def test_find_header_row_detects_later_row(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "inventory.xlsx",
        [
            ["junk", "", ""],
            ["PART NUMBER", "SUB-SYSTEM", "QPA", "SERIAL NUMBER"],
            ["ABC-123", "RADAR", 2, "S1"],
        ],
    )
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row_idx, headers = find_header_row(ws)
    wb.close()
    assert row_idx == 2
    assert "part_number" in headers


def test_remap_corpus_path_returns_existing(tmp_path: Path):
    path = tmp_path / "exists.xlsx"
    path.write_text("x")
    assert remap_corpus_path(path) == path


def test_parse_site_inventory_workbook(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Site Inventory Report_Guam.xlsx",
        [
            ["PART NUMBER", "SUB-SYSTEM", "NOMENCLATURE", "QPA", "MODEL NUMBER", "SERIAL NUMBER"],
            ["ARC-4471", "RADAR", "Widget", 2, "M1", "SN1234"],
        ],
        sheet_name="GUAM SITE INVENTORY REPORT",
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Guam\2024-03-01\Site Inventory Report_Guam.xlsx", system="NEXION", site_token="guam")
    assert len(rows) == 1
    assert rows[0].part_number == "ARC-4471"
    assert rows[0].quantity_at_site == 2
    assert rows[0].site_token == "guam"
    assert rows[0].extraction_method == "xlsx_site_inventory_qpa_v1"


def test_parse_inventory_spares_workbook(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Inventory & Spares.xlsx",
        [
            ["SUB-SYSTEM", "ITEM TYPE", "PART NUMBER", "SERIAL NUMBER", "QTY"],
            ["RF", "INSTALLED", "SEMS3D-40536", "", 3],
        ],
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Learmonth\2024-06-07\Inventory & Spares.xlsx", system="NEXION", site_token="learmonth")
    assert len(rows) == 1
    assert rows[0].part_number == "SEMS3D-40536"
    assert rows[0].quantity_at_site == 3
    assert rows[0].extraction_method == "xlsx_inventory_spares_qty_v1"


def test_parse_as_built_parts_list_workbook(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Wake Parts List.xlsx",
        [
            ["FIND NO.", "QTY.", "PART NUMBER", "DESCRIPTION", "MANUFACTURER", "SERIAL NO."],
            [1, 4, "SEMS3D-40540", "Panel", "NG", "AB1234"],
        ],
        sheet_name="Parts List",
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Wake\2024-12-02\Wake Parts List.xlsx", system="NEXION", site_token="wake")
    assert len(rows) == 1
    assert rows[0].part_number == "SEMS3D-40540"
    assert rows[0].quantity_at_site == 4
    assert rows[0].serial_number == "AB1234"
    assert rows[0].extraction_method == "xlsx_as_built_parts_list_v1"


def test_skip_packing_list_sheet(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "NG Packing List.xlsx",
        [
            ["PART NUMBER", "QTY", "SERIAL NUMBER"],
            ["ARC-4471", 2, "SN1"],
        ],
        sheet_name="Packing List",
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\ISTO\Ascension\2024-03-01\NG Packing List.xlsx", system="ISTO", site_token="ascension")
    assert rows == []


def test_skip_cable_table_sheet(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Cable Table.xlsx",
        [
            ["EQUIP.", "SOCKET", "FROM", "TO"],
            ["A", "J1", "P1", "P2"],
        ],
        sheet_name="Cable Table",
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Guam\2024-03-01\Cable Table.xlsx", system="NEXION", site_token="guam")
    assert rows == []


def test_serial_fallback_implies_quantity_one(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Parts List.xlsx",
        [
            ["QTY.", "PART NUMBER", "DESCRIPTION", "MANUFACTURER", "SERIAL NO."],
            ["", "ARC-4471", "Widget", "NG", "SER1001"],
        ],
        sheet_name="Parts List",
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Guam\2024-03-01\Parts List.xlsx", system="NEXION", site_token="guam")
    assert len(rows) == 1
    assert rows[0].quantity_at_site == 1


def test_duplicate_rows_deduped_within_sheet(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Inventory.xlsx",
        [
            ["PART NUMBER", "QPA", "SERIAL NUMBER"],
            ["ARC-4471", 2, "SN1"],
            ["ARC-4471", 2, "SN1"],
        ],
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Guam\2024-03-01\Inventory.xlsx", system="NEXION", site_token="guam")
    assert len(rows) == 1


def test_missing_part_number_skipped(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Inventory.xlsx",
        [
            ["PART NUMBER", "QPA", "SERIAL NUMBER"],
            ["", 2, "SN1"],
        ],
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Guam\2024-03-01\Inventory.xlsx", system="NEXION", site_token="guam")
    assert rows == []


def test_invalid_quantity_without_serial_skipped(tmp_path: Path):
    path = _write_workbook(
        tmp_path / "Inventory.xlsx",
        [
            ["PART NUMBER", "QPA", "SERIAL NUMBER"],
            ["ARC-4471", "", ""],
        ],
    )
    rows = parse_installed_base_workbook(path, source_path=r"D:\CorpusTransfr\NEXION\Guam\2024-03-01\Inventory.xlsx", system="NEXION", site_token="guam")
    assert rows == []


def test_content_derived_fallback_still_extracts_from_tmp_path(tmp_path: Path):
    moved = _write_workbook(
        tmp_path / "moved.xlsx",
        [
            ["PART NUMBER", "QTY", "SERIAL NUMBER"],
            ["WR-200", 5, ""],
        ],
    )
    rows = parse_installed_base_workbook(
        moved,
        source_path=str(moved),
        system="NEXION",
        site_token="wake",
    )
    assert len(rows) == 1
    assert rows[0].part_number == "WR-200"
    assert rows[0].quantity_at_site == 5
