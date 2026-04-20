"""Workbook parser for installed-base denominator families.

v1 scope is intentionally narrow:
  - Site Inventory Report workbooks (QPA template)
  - Inventory & Spares / Pre-inventory workbooks (QTY template)
  - As-Built Parts List workbooks (QTY + PART NUMBER template)
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from src.extraction.failure_event_extractor import detect_system, extract_part_numbers, extract_year
from src.extraction.installed_base_extractor import extract_snapshot_date
from src.store.installed_base_store import InstalledBaseRecord
from src.vocab.pack_loader import resolve_column

logger = logging.getLogger(__name__)

_ACTIVE_PATH_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"site[\s_-]*inventory.*\.xlsx$", re.IGNORECASE),
    re.compile(r"inventory.*spares?.*\.xlsx$", re.IGNORECASE),
    re.compile(r"pre[\s_-]*inventory.*\.xlsx$", re.IGNORECASE),
    re.compile(r"as[\s_-]*built.*\.xlsx$", re.IGNORECASE),
)

_SKIP_SHEET_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"packing\s+list", re.IGNORECASE),
    re.compile(r"cable\s+table", re.IGNORECASE),
    re.compile(r"miscellaneous", re.IGNORECASE),
)

_SITE_TOKEN_RE = re.compile(
    r"\b(alpena|ascension|awase|azores|curacao|diego\s+garcia|djibouti|eglin|fairford|"
    r"guam|hawaii|kwajalein|learmonth|lualualei|misawa|niger|okinawa|palau|"
    r"pituffik|thule|vandenberg|wake)\b",
    re.IGNORECASE,
)


def is_supported_installed_base_xlsx(source_path: str | Path) -> bool:
    path_text = str(source_path or "")
    return path_text.lower().endswith(".xlsx") and any(p.search(path_text) for p in _ACTIVE_PATH_PATTERNS)


def remap_corpus_path(source_path: str | Path) -> Path:
    original = Path(str(source_path))
    if original.exists():
        return original
    path_str = str(source_path)
    if path_str.startswith("D:\\CorpusTransfr"):
        remapped = Path(path_str.replace("D:\\CorpusTransfr", "E:\\CorpusTransfr", 1))
        if remapped.exists():
            return remapped
    return original


def normalize_header(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.strip().lower() == "actual serial number":
        return "actual_serial_number"
    for concept in ("part_number", "quantity", "serial_number"):
        resolved = resolve_column(text, concept)
        if resolved:
            return resolved
    return text.strip().lower().replace(" ", "_")


def _normalize_cell_text(value: object) -> str:
    return str(value or "").strip()


def detect_template(normalized_headers: list[str]) -> str:
    header_set = {h for h in normalized_headers if h}
    if "part_number" not in header_set or "quantity" not in header_set:
        return ""
    if {"description", "manufacturer"} & header_set:
        return "as_built_parts_list"
    if "item_type" in header_set:
        return "inventory_spares_qty"
    return "site_inventory_qpa"


def find_header_row(worksheet, max_scan_rows: int = 8) -> tuple[int | None, list[str]]:
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        normalized = [normalize_header(cell) for cell in row]
        if detect_template(normalized):
            return row_idx, normalized
    return None, []


def derive_site_token(source_path: str, fallback_site_token: str = "") -> str:
    if fallback_site_token:
        return fallback_site_token.lower()
    match = _SITE_TOKEN_RE.search(source_path or "")
    if not match:
        return ""
    token = match.group(1).lower()
    return "thule" if token == "pituffik" else token


def derive_snapshot_year(source_path: str, snapshot_date: str) -> int | None:
    return extract_year(snapshot_date) or extract_year(source_path)


def _coerce_quantity(value: object) -> int | None:
    text = _normalize_cell_text(value)
    if not text:
        return None
    text = text.replace(",", "")
    try:
        if "." in text:
            number = int(float(text))
        else:
            number = int(text)
    except ValueError:
        return None
    return number if 0 < number < 100000 else None


def _extract_primary_part_number(raw_part: str) -> str:
    parts = extract_part_numbers(raw_part, max_parts=1)
    if parts:
        return parts[0]
    token = raw_part.strip().upper().replace(" ", "")
    if re.fullmatch(r"[A-Z0-9][A-Z0-9._/-]{2,}", token):
        return token
    return ""


def _looks_like_skip_sheet(sheet_name: str, normalized_headers: list[str]) -> bool:
    if any(p.search(sheet_name or "") for p in _SKIP_SHEET_PATTERNS):
        return True
    header_set = {h for h in normalized_headers if h}
    if {"socket", "socket_type", "from", "to"} & header_set:
        return True
    if {"sheet_number", "drawing_title"} <= header_set:
        return True
    return False


def parse_installed_base_workbook(
    workbook_path: str | Path,
    *,
    source_path: str | None = None,
    system: str = "",
    site_token: str = "",
    source_doc_hash: str = "",
) -> list[InstalledBaseRecord]:
    resolved_path = remap_corpus_path(workbook_path)
    if not resolved_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    source_path = str(source_path or workbook_path)
    system = system or detect_system(source_path) or detect_system(str(resolved_path))
    site_token = derive_site_token(source_path, site_token)
    snapshot_date = extract_snapshot_date(source_path) or extract_snapshot_date(str(resolved_path))
    snapshot_year = derive_snapshot_year(source_path, snapshot_date)

    wb = load_workbook(resolved_path, read_only=True, data_only=True)
    records: list[InstalledBaseRecord] = []
    seen: set[tuple[str, str, str]] = set()

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_idx, normalized_headers = find_header_row(ws)
            if header_row_idx is None or _looks_like_skip_sheet(sheet_name, normalized_headers):
                continue
            template = detect_template(normalized_headers)
            if not template:
                continue
            header_map = {
                normalized: idx
                for idx, normalized in enumerate(normalized_headers)
                if normalized
            }
            for row_offset, row in enumerate(
                ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
                start=header_row_idx + 1,
            ):
                part_raw = _normalize_cell_text(row[header_map["part_number"]]) if "part_number" in header_map else ""
                if not part_raw:
                    continue
                part_number = _extract_primary_part_number(part_raw)
                if not part_number:
                    continue

                quantity = None
                if "quantity" in header_map:
                    quantity = _coerce_quantity(row[header_map["quantity"]])

                serial = ""
                for serial_key in ("actual_serial_number", "serial_number"):
                    if serial_key in header_map:
                        serial = _normalize_cell_text(row[header_map[serial_key]])
                        if serial:
                            break

                if quantity is None and serial:
                    quantity = 1
                if quantity is None:
                    continue

                dedupe_key = (sheet_name, part_number, serial or f"qty:{quantity}")
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)

                records.append(
                    InstalledBaseRecord(
                        source_path=source_path,
                        source_doc_hash=source_doc_hash,
                        chunk_id=f"xlsx:{sheet_name}:{row_offset}",
                        part_number=part_number,
                        serial_number=serial,
                        system=system,
                        site_token=site_token,
                        install_date="",
                        snapshot_date=snapshot_date,
                        snapshot_year=snapshot_year,
                        quantity_at_site=quantity,
                        extraction_method=f"xlsx_{template}_v1",
                        confidence=0.93 if quantity is not None else 0.85,
                    )
                )
    finally:
        wb.close()

    return records


def iter_xlsx_source_candidates(metadata_db: str | Path, *, limit: int | None = None) -> Iterable[dict[str, str]]:
    import sqlite3

    conn = sqlite3.connect(str(metadata_db))
    try:
        sql = """
            SELECT source_path, site_token, source_doc_hash
            FROM source_metadata
            WHERE lower(source_path) LIKE '%.xlsx'
        """
        count = 0
        for source_path, site_token, source_doc_hash in conn.execute(sql):
            if not is_supported_installed_base_xlsx(source_path):
                continue
            yield {
                "source_path": str(source_path or ""),
                "site_token": str(site_token or ""),
                "source_doc_hash": str(source_doc_hash or ""),
            }
            count += 1
            if limit is not None and count >= int(limit):
                return
    finally:
        conn.close()


def parse_candidate_workbooks(metadata_db: str | Path, *, limit: int | None = None) -> Iterable[InstalledBaseRecord]:
    for item in iter_xlsx_source_candidates(metadata_db, limit=limit):
        source_path = item["source_path"]
        system = detect_system(source_path)
        if not system:
            continue
        try:
            rows = parse_installed_base_workbook(
                source_path,
                source_path=source_path,
                system=system,
                site_token=item["site_token"],
                source_doc_hash=item["source_doc_hash"],
            )
        except Exception as exc:
            logger.warning("installed_base_xlsx parse failed for %s: %s", source_path, exc)
            continue
        for row in rows:
            yield row
